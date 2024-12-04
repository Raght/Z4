from package import *
from typing import Union
import numpy
import matplotlib.pyplot as plt
import openpyxl
import glob


def print_reservuar_private_members(container: Reservuar, class_name: str):
    members = ("__HH", "__RR", "__FF")
    for i in range(len(members) - 1):
        attribute_name = "_" + class_name + members[i]
        value = container.__dict__[attribute_name]
        print(f"'{attribute_name}' = {value}, ", end="")
    attribute_name = "_" + class_name + members[len(members) - 1]
    value = container.__dict__[attribute_name]
    print(f"'{attribute_name}' = {value}")

def print_private_members_of_containers(containers: Union[tuple[ReservuarShestiugolnikTreugolnik], list[ReservuarShestiugolnikTreugolnik]]):
    for container in containers:
        print_reservuar_private_members(container, "ReservuarShestiugolnikTreugolnik")

#def next_h(volume: float, function):
#    h0 = math.cbrt(volume)
#    h = h0 / 3
#    n = 50
#    for i in range(n // 2):
#        h = h + (h0 - h) * i / (n // 2)
#        yield (h, function(h))
#    yield (h0, function(h0))
#    h_max = h0 * 2
#    for i in range(n // 2 + 1, n):
#        h = h + (h_max - h0) * (i - n // 2 - 1) / (n - 1)
#        yield (h, function(h))


N = 50


#def next_h(volume: float, function):
#    h0 = math.cbrt(volume)
#    yield (h0, function(h0))
#    h = h0
#    for i in range(N):
#        h /= 3
#        yield (h, function(h))
#    h_max = h0 * 2
#    for i in range(1, N + 1):
#        h = h + (h_max - h0) * i / N
#        yield (h, function(h))



def next_h(volume: float, function):
    global N
    h0 = math.cbrt(volume)
    h_min = h0 / 3
    for i in range(N):
        h = h_min + (h0 - h_min) * i / N
        yield (h, function(h))
    yield (h0, function(h0))
    h_max = h0 * 2
    for i in range(N):
        h = h0 + (h_max - h0) * i / (N - 1)
        yield (h, function(h))


def build_and_save_plot(volume, function, y_name: str, y_label: str, title: str, show_plot: bool = True, filename_jpg: str = "", filename_txt: str = "", filename_xlsx: str = "", filename_csv: str = ""):
    pairs = []
    get_next_h = next_h(volume, function)
    while True:
        try:
            pairs.append(next(get_next_h))
        except StopIteration:
            break
    h = [pairs[i][0] for i in range(len(pairs))]
    y = [pairs[i][1] for i in range(len(pairs))]

    x_name = "h"
    x_label = "h - height"

    plt.plot(h, y)
    plt.xlabel(x_label)
    plt.ylabel(y_label)
    plt.title(title)

    if show_plot:
        plt.show()

    if filename_jpg != "":
        plt.savefig(filename_jpg)


    if filename_txt != "":
        rows: list[list[str]] = [[] for _ in range(len(pairs) + 1)]
        lengths: list[int] = [0 for _ in range(len(pairs) + 1)]
        variable_names = [x_name, y_name]

        with open(filename_txt, "w") as file:
            for i in range(2):
                rows[0].append(variable_names[i])
                lengths[0] += len(variable_names[i])
                for i in range(1, len(pairs) + 1):
                    h_str = str(h[i-1])
                    rows[i].append(h_str)
                    lengths[i] += len(h_str)

                max_length = 0
                for i in range(len(pairs) + 1):
                    max_length = max(max_length, lengths[i])

                for i in range(len(pairs) + 1):
                    spaces_str_length = max_length - lengths[i] + 1
                    rows[i].append(" " * spaces_str_length)
                    lengths[i] += spaces_str_length

            for row in rows:
                file.write("".join(row))
                file.write("\n")

    if filename_csv != "":
        with open(filename_csv, "w") as file:
            file.write(f"{x_name},{y_name}")
            file.write("\n")
            for i in range(len(pairs)):
                file.write(f"{h[i]},{y[i]}")
                file.write("\n")

    if filename_xlsx != "":
        workbook = openpyxl.Workbook()
        sheet_name = "Лист 1"
        workbook.create_sheet(sheet_name)
        if glob.glob(filename_xlsx):
            workbook = openpyxl.load_workbook(filename_xlsx)
        sheet = workbook[sheet_name]
        sheet.cell(row=1, column=1, value=x_name)
        sheet.cell(row=1, column=2, value=y_name)
        for i in range(len(pairs)):
            sheet.cell(row=i+2, column=1, value=str(h[i]))
            sheet.cell(row=i+2, column=2, value=str(y[i]))
        workbook.save(filename_xlsx)


container1 = ReservuarShestiugolnikTreugolnik(Material.STEEL, 150, 0.25)
container2 = ReservuarShestiugolnikTreugolnik(Material.BRASS, 175, 0.3)
container3 = ReservuarShestiugolnikTreugolnik(Material.ALUMINIUM_ALLOY, 125, 0.15)
container4 = ReservuarShestiugolnikTreugolnik(Material.TITANIUM_ALLOY, 200, 0.175)
container5 = ReservuarShestiugolnikTreugolnik(Material.POLYMER_COMPOSITE, 250, 0.35)

containers = [container1, container2, container3, container4, container5]

for container in containers:
    print(container.__dict__)
for i in range(len(containers)):
    print(f"container{i + 1}: ")
    print_reservuar_private_members(containers[i], "Reservuar")
print("Performing optimization")
for container in containers:
    container.optimize()
for container in containers:
    print(container.__dict__)
for i in range(len(containers)):
    print(f"container{i + 1}: ")
    print_reservuar_private_members(containers[i], "ReservuarShestiugolnikTreugolnik")

print("Sorting by volume")
containers.sort(key=lambda container: container.volume)
print_private_members_of_containers(containers)
print("Sorting by coefficient c")
containers.sort(key=lambda container: container.c_coefficient)
print_private_members_of_containers(containers)
print("Sorting by material")
containers.sort(key=lambda container: container.material)
print_private_members_of_containers(containers)
print("Sorting by surface area")
containers.sort(key=lambda container: container.surface_area)
print_private_members_of_containers(containers)
print("Sorting by outer size R")
containers.sort(key=lambda container: container.outer_size)
print_private_members_of_containers(containers)
print("Sorting by height H")
containers.sort(key=lambda container: container.height)
print_private_members_of_containers(containers)

container010 = ReservuarShestiugolnikTreugolnik(Material.STEEL, 150, 0.25)
container015 = ReservuarShestiugolnikTreugolnik(Material.STEEL, 200, 0.15)
container020 = ReservuarShestiugolnikTreugolnik(Material.STEEL, 200, 0.2)
container025 = ReservuarShestiugolnikTreugolnik(Material.STEEL, 200, 0.25)

build_and_save_plot(container010.volume, container010.compute_outer_size_R, "R","R - outer size", "R = R(h)",
                    filename_jpg="R, c=0.1.jpg",
                    filename_txt="R, c=0.1.txt",
                    filename_xlsx="R, c=0.1.xlsx",
                    filename_csv="R, c=0.1.csv")
build_and_save_plot(container010.volume, container010.compute_area_from_height, "F", "F - surface area", "F = F(h)",
                    filename_jpg="F, c=0.1.jpg",
                    filename_txt="F, c=0.1.txt",
                    filename_xlsx="F, c=0.1.xlsx",
                    filename_csv="F, c=0.1.csv")
build_and_save_plot(container015.volume, container015.compute_outer_size_R, "R", "R - outer size", "R = R(h)",
                    filename_jpg="R, c=0.15.jpg",
                    filename_txt="R, c=0.15.txt",
                    filename_xlsx="",
                    filename_csv="")
build_and_save_plot(container015.volume, container015.compute_area_from_height, "F", "F - surface area", "F = F(h)",
                    filename_jpg="F, c=0.15.jpg",
                    filename_txt="F, c=0.15.txt",
                    filename_xlsx="",
                    filename_csv="")
build_and_save_plot(container020.volume, container020.compute_outer_size_R, "R", "R - outer size", "R = R(h)",
                    filename_jpg="R, c=0.2.jpg",
                    filename_txt="R, c=0.2.txt",
                    filename_xlsx="",
                    filename_csv="",
                    show_plot=False)
build_and_save_plot(container020.volume, container020.compute_area_from_height, "F", "F - surface area", "F = F(h)",
                    filename_jpg="F, c=0.2.jpg",
                    filename_txt="F, c=0.2.txt",
                    filename_xlsx="",
                    filename_csv="",
                    show_plot=False)
build_and_save_plot(container025.volume, container025.compute_outer_size_R, "R", "R - outer size", "R = R(h)",
                    filename_jpg="R, c=0.25.jpg",
                    filename_txt="R, c=0.25.txt",
                    filename_xlsx="",
                    filename_csv="",
                    show_plot=False)
build_and_save_plot(container025.volume, container025.compute_area_from_height, "F", "F - surface area", "F = F(h)",
                    filename_jpg="F, c=0.25.jpg",
                    filename_txt="F, c=0.25.txt",
                    filename_xlsx="",
                    filename_csv="",
                    show_plot=False)

